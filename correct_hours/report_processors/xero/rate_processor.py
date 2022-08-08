import dataclasses
from datetime import datetime
from typing import Optional

from openpyxl import Workbook

from correct_hours.report_processors.types import RateNotFound

ROWS_OFFSET = 2
COLS_OFFSET = 3


@dataclasses.dataclass
class Rate:
    label: str
    start_date: Optional[datetime]
    end_date: Optional[datetime]
    rate: float

    def is_within_date_range(self, date: datetime) -> bool:
        if self.start_date is None:
            return date < self.end_date
        if self.end_date is None:
            return date >= self.start_date
        return self.start_date <= date < self.end_date


class XeroRateProcessor:

    def __init__(self, workbook: Workbook) -> None:
        self.workbook = workbook
        self.rates = []

    def process(self) -> None:
        return self.process_rows()

    def process_rows(self) -> None:
        sheet = self.workbook.active
        rates = []
        for row in sheet.iter_rows(min_row=ROWS_OFFSET, values_only=True):
            rate_label = row[0]
            for col_idx, col in enumerate(row):
                col_number = col_idx + 1
                if col_number < COLS_OFFSET:
                    # skip first columns
                    continue
                elif col_number == COLS_OFFSET:
                    # oldest rate
                    start_date = None
                    end_date = sheet.cell(1, col_number + 1).value
                elif col_idx >= len(row):
                    # latest rate
                    start_date = sheet.cell(1, col_number).value
                    end_date = None
                else:
                    start_date = sheet.cell(1, col_number).value
                    end_date = sheet.cell(1, col_number + 1).value
                rate = col
                rates.append(
                    Rate(rate_label, start_date, end_date, rate)
                )
        self.rates = rates

    def get_rate(self, rate_label: str, date: datetime) -> int:
        matching_rates = [
            r.rate for r in self.rates if str.lower(r.label) == str.lower(rate_label) and r.is_within_date_range(date)
        ]
        if matching_rates:
            return matching_rates[0]
        else:
            raise RateNotFound(rate_label, date)




