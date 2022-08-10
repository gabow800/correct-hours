from pathlib import Path
from typing import Tuple, List

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from correct_hours.report_processors.xero import XeroReportProcessor, ROW_OFFSET
from correct_hours.utils import get_col_number, get_col_name

HOUR_START_ROW = 6


def assert_other_cells_should_be_the_same(
        old_sheet: Worksheet,
        new_sheet: Worksheet,
        skip_cells: List[str]
) -> None:
    for row_idx, row in enumerate(new_sheet.iter_rows(min_row=ROW_OFFSET, values_only=True)):
        row_number = row_idx + ROW_OFFSET
        if row_number == new_sheet.max_row:
            # skip bottom row
            continue
        for col_number in range(get_col_number("A"), get_col_number("N") + 1):
            col_name = get_col_name(col_number)
            cell_name = f"{col_name}{row_number}"
            if cell_name in skip_cells:
                continue
            old_value = old_sheet[cell_name].value
            new_value = new_sheet[cell_name].value
            assert old_value == new_value, (
                f"{cell_name} should have the same values in both sheets: {old_value} == {new_value}"
            )


def assert_corrected_hours(
        old_sheet: Worksheet,
        new_sheet: Worksheet,
        corrected_cells: List[Tuple[str, int, int]]
) -> None:
    for corrected_cell in corrected_cells:
        cell_name, expected_old_value, expected_new_value = corrected_cell
        old_value = old_sheet[cell_name].value
        new_value = new_sheet[cell_name].value
        assert old_value != new_value
        assert old_value == expected_old_value
        assert new_value == expected_new_value


def assert_cell_values(new_sheet: Worksheet, cells: List[Tuple[str, str]]) -> None:
    for cell in cells:
        cell_name, expected_value = cell
        assert new_sheet[cell_name].value == expected_value


def test_process_file() -> None:
    hours_workbook = load_workbook("tests/data/xero-report.xlsx")
    rates_workbook = load_workbook("tests/data/rates.xlsx")
    processor = XeroReportProcessor(hours_workbook, rates_workbook)
    processor.process()
    Path(f"tests/data/output").mkdir(parents=True, exist_ok=True)
    hours_workbook.save(filename="tests/data/output/copy-xero-report.xlsx")
    old_sheet = hours_workbook["Timesheet Details"]
    new_sheet = hours_workbook["Timesheet Details Copy"]
    # Asert old total, new total and days worked
    assert_cell_values(new_sheet, [
        # 1/12/2020 rows
        ("O6", "=SUM(G6:M6)"),
        ("O7", "=SUM(G7:M7)"),
        ("O8", "=SUM(G8:M8)"),
        ("P8", "=SUM(O6:O8)"),
        ("Q8", "=P8/7.6"),
        # 2/12/2020 rows
        ("O9", "=SUM(G9:M9)"),
        ("P9", "=SUM(O9:O9)"),
        ("Q9", "=P9/7.6"),
        # 3/12/2020 rows
        ("O10", "=SUM(G10:M10)"),
        ("O11", "=SUM(G11:M11)"),
        ("P11", "=SUM(O10:O11)"),
        ("Q11", "=P11/7.6"),
        # 4/12/2020 rows
        ("O12", "=SUM(G12:M12)"),
        ("P12", "=SUM(O12:O12)"),
        ("Q12", "=P12/7.6"),
    ])
    # Assert corrected hours process
    corrected_hours = [
        ("I8", 2, 0),
        ("I7", 2, 0),
        ("I6", 2, 0),
        ("H8", 2, 0),
        ("H7", 7, 6),
        ("K12", 8, 6),
    ]
    assert_corrected_hours(old_sheet, new_sheet, corrected_hours)
    # Other cells should be the same
    skip_cells = [corrected_hour[0] for corrected_hour in corrected_hours]
    assert_other_cells_should_be_the_same(
        old_sheet,
        new_sheet,
        skip_cells
    )

    # Assert that rates have been applied
    assert_cell_values(new_sheet, [
        ("R6", "=O6*25.52"),
        ("R7", "=O7*36.24"),
        ("R8", "=O8*48.32"),
        ("R9", "=O9*25.52"),
        ("R10", "=O10*25.52"),
        ("R11", "=O11*36.24"),
        ("R12", "=O12*25.52"),
   ])

