from pathlib import Path

from openpyxl import load_workbook

from correct_hours.report_processors.xero import XeroReportProcessor

HOUR_COLUMN_NAMES = ["G", "H", "I", "J", "K", "L", "M", "N"]
HOUR_START_ROW = 6


def assert_other_hours_should_be_the_same(old_sheet, new_sheet, skip_cells):
    for col_name in HOUR_COLUMN_NAMES:
        for row_idx, row in enumerate(new_sheet.iter_rows(min_row=HOUR_START_ROW, values_only=True)):
            row_number = row_idx + HOUR_START_ROW
            cell_name = f"{col_name}{row_number}"
            if cell_name in skip_cells:
                continue
            old_value = old_sheet[cell_name].value
            new_value = new_sheet[cell_name].value
            assert old_value == new_value, (
                f"{cell_name} should have the same values in both sheets: {old_value} == {new_value}"
            )


def assert_corrected_hours(old_sheet, new_sheet, corrected_cells):
    for corrected_cell in corrected_cells:
        cell_name, expected_old_value, expected_new_value = corrected_cell
        old_value = old_sheet[cell_name].value
        new_value = new_sheet[cell_name].value
        assert old_value != new_value
        assert old_value == expected_old_value
        assert new_value == expected_new_value


def assert_cell_values(new_sheet, cells):
    for cell in cells:
        cell_name, expected_value = cell
        assert new_sheet[cell_name].value == expected_value


def test_process_file():
    workbook = load_workbook("tests/data/timesheet-david.xlsx")
    processor = XeroReportProcessor(workbook)
    processor.process()
    Path(f"tests/data/output").mkdir(parents=True, exist_ok=True)
    workbook.save(filename="tests/data/output/copy-timesheet-david.xlsx")
    old_sheet = workbook.get_sheet_by_name("Timesheet Details")
    new_sheet = workbook.get_sheet_by_name("Timesheet Details Copy")
    # Asert old total, new total and days worked
    assert_cell_values(new_sheet, [
        # 1/12/2020 rows
        ("O8", 47),
        ("P8", "=SUM(G6:M8)"),
        ("Q8", "=P8/7.6"),
        # 2/12/2020 rows
        ("O9", 32),
        ("P9", "=SUM(G9:M9)"),
        ("Q9", "=P9/7.6"),
        # 3/12/2020 rows
        ("O11", 34),
        ("P11", "=SUM(G10:M11)"),
        ("Q11", "=P11/7.6"),
        # 4/12/2020 rows
        ("O12", 40),
        ("P12", "=SUM(G12:M12)"),
        ("Q12", "=P12/7.6"),
    ])
    # Assert corrected hours process
    corrected_hours = [
        ("I8", 2, 0),
        ("I7", 2, 0),
        ("I6", 2, 0),
        ("H8", 2, 0),
        ("H7", 7, 6),
        ("K12", 8, 6),
    ]
    assert_corrected_hours(old_sheet, new_sheet, corrected_hours)
    # Other hours should be the same
    assert_other_hours_should_be_the_same(
        old_sheet,
        new_sheet,
        [corrected_hour[0] for corrected_hour in corrected_hours]
    )

